#!/usr/bin/env python3
"""한국투자증권 OpenAPI 엑셀 (339시트) → LLM 친화적 markdown 변환.

엑셀은 한 시트당 1 API이며 행/열로 메타데이터 + Layout(요청/응답) 표가 들어 있다.
LLM은 (1) cp949 인코딩, (2) 셀 단위 구조, (3) 339개 시트 검색이 어렵다.

본 스크립트는:
  - docs/kis-api/INDEX.md      — 전체 API 목록 (TR_ID·카테고리·파일·앵커)
  - docs/kis-api/<category>.md — 카테고리별 섹션 합본 (5개 + 기타)
  - 각 API당 1 섹션: 메타 헤더 + Request/Response 표 + 예시
  - OAuth 공통 헤더(authorization, appkey 등)는 중복 제거 (INDEX.md 상단에 한 번만)

Usage: python scripts/build-api-docs.py
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "한국투자증권_오픈API_전체문서_20260418_030007.xlsx"
OUT = ROOT / "docs" / "kis-api"

# 메뉴 위치 prefix → 파일명
CATEGORY_FILE = {
    "[국내주식]": "domestic-stock.md",
    "[해외주식]": "overseas-stock.md",
    "[국내선물옵션]": "domestic-futureoption.md",
    "[해외선물옵션]": "overseas-futureoption.md",
    "[장내채권]": "domestic-bond.md",
}
OTHER_FILE = "_other.md"

# 모든 API 공통이라 시트별 중복 출력 제외할 헤더 Element
COMMON_HEADER_ELEMENTS = {
    "content-type", "authorization", "appkey", "appsecret", "personalseckey",
    "tr_id", "tr_cont", "custtype", "seq_no", "mac_address", "phone_number",
    "ip_addr", "gt_uid",
}

META_LABELS = {
    "API명": "name",
    "API 명": "name",
    "API ID": "api_id",
    "실전 TR_ID": "tr_real",
    "모의 TR_ID": "tr_paper",
    "HTTP Method": "method",
    "URL 명": "url",
    "메뉴 위치": "menu",
    "설명": "desc",
    "실전 Domain": "domain_real",
    "모의 Domain": "domain_paper",
    "통신구분": "transport",
    "API 통신구분": "transport",
    "기본정보": "_skip",
    "Layout": "_skip",
}


def normalize_cell(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # 줄바꿈은 markdown table에서 깨지므로 한 줄로
    return s.replace("\r", " ").replace("\n", " ").replace("|", "\\|")


def slug(s: str) -> str:
    """시트명 → 안정적인 markdown anchor."""
    s = re.sub(r"[^\w가-힣()_-]+", "-", s.strip().lower())
    s = re.sub(r"-+", "-", s).strip("-")
    return s or "api"


def extract_sheet(name: str, sheet) -> dict:
    """엑셀 시트 한 장 → dict {meta, layout: [(section, element, kor, type, req, len, desc)], examples: [...]}"""
    rows = list(sheet.iter_rows(values_only=True))
    meta: dict[str, str] = {"sheet": name}
    layout: list[tuple[str, str, str, str, str, str, str]] = []
    request_example = ""
    response_example = ""

    layout_started = False
    current_section = ""
    in_response_example = False
    response_example_buffer: list[str] = []

    for r in rows:
        if not r:
            continue
        cells = [normalize_cell(c) for c in r]
        # 빈 행이면 응답 예시 수집 끊기
        if not any(cells):
            if in_response_example:
                in_response_example = False
            continue

        first = cells[0]

        # 응답 예시 수집 모드
        if in_response_example:
            joined = " | ".join(c for c in cells if c)
            if joined:
                response_example_buffer.append(joined)
            continue

        # 메타데이터 (R1~R14 부근)
        if first in META_LABELS:
            key = META_LABELS[first]
            if key != "_skip" and len(cells) > 1:
                meta[key] = cells[1]
            if first == "Layout":
                layout_started = True
            continue

        # Layout 영역
        if layout_started:
            # 헤더 행 ("구분|Element|한글명|Type|Required|Length|Description")
            if first == "구분" and len(cells) >= 7:
                continue

            # Request Example
            if "Request Example" in first:
                # cells[1]에 example 본문이 들어 있는 경우가 많음
                if len(cells) > 1:
                    request_example = cells[1]
                continue
            if "Response Example" in first:
                if len(cells) > 1 and cells[1].strip():
                    response_example_buffer.append(cells[1])
                in_response_example = True
                continue

            # 일반 layout 행: (구분, Element, 한글명, Type, Required, Length, Description)
            # 구분 칸이 비어 있으면 직전 섹션 유지
            section = first or current_section
            if section:
                current_section = section
            element = cells[1] if len(cells) > 1 else ""
            kor = cells[2] if len(cells) > 2 else ""
            ty = cells[3] if len(cells) > 3 else ""
            req = cells[4] if len(cells) > 4 else ""
            ln = cells[5] if len(cells) > 5 else ""
            desc = cells[6] if len(cells) > 6 else ""
            if element:
                layout.append((current_section, element, kor, ty, req, ln, desc))

    response_example = "\n".join(response_example_buffer)
    return {
        "meta": meta,
        "layout": layout,
        "request_example": request_example,
        "response_example": response_example,
    }


def categorize(meta: dict) -> tuple[str, str]:
    """반환 (file_name, category_label)."""
    menu = meta.get("menu", "")
    for prefix, fname in CATEGORY_FILE.items():
        if menu.startswith(prefix):
            return fname, prefix
    return OTHER_FILE, "(기타)"


def render_section(api: dict, idx: int) -> str:
    meta = api["meta"]
    layout = api["layout"]
    name = meta.get("name") or meta.get("sheet", "")
    sheet = meta["sheet"]
    tr_real = meta.get("tr_real", "").strip()
    tr_paper = meta.get("tr_paper", "").strip()
    method = meta.get("method", "GET").strip() or "GET"
    url = meta.get("url", "").strip()
    menu = meta.get("menu", "").strip()
    desc = meta.get("desc", "").strip()

    parts: list[str] = []
    parts.append(f"### {idx}. {name}")
    parts.append("")
    # 메타 표
    meta_rows = [
        ("Sheet", f"`{sheet}`"),
        ("Menu", menu or "—"),
        ("Method", f"`{method}`"),
        ("URL", f"`{url}`" if url else "—"),
        ("TR_ID (실전)", f"`{tr_real}`" if tr_real else "—"),
    ]
    if tr_paper and tr_paper != tr_real and "동일" not in tr_paper:
        meta_rows.append(("TR_ID (모의)", f"`{tr_paper}`"))
    parts.append("| Field | Value |")
    parts.append("|---|---|")
    for k, v in meta_rows:
        parts.append(f"| {k} | {v} |")
    parts.append("")

    if desc:
        parts.append("**설명:** " + desc)
        parts.append("")

    # Layout 섹션별 그룹핑
    sections: dict[str, list[tuple]] = {}
    section_order: list[str] = []
    for sec, *rest in layout:
        # 공통 OAuth 헤더는 응답에서도 비슷하니 Request Header에서만 필터
        if sec.startswith("Request Header") and rest[0].lower() in COMMON_HEADER_ELEMENTS:
            continue
        if sec.startswith("Response Header") and rest[0].lower() in {"content-type", "tr_id", "tr_cont", "gt_uid"}:
            continue
        sections.setdefault(sec, []).append(tuple(rest))
        if sec not in section_order:
            section_order.append(sec)

    for sec in section_order:
        rows = sections[sec]
        if not rows:
            continue
        parts.append(f"#### {sec}")
        parts.append("")
        parts.append("| Element | 한글명 | Type | Req | Len | Description |")
        parts.append("|---|---|---|---|---|---|")
        for el, kor, ty, req, ln, ds in rows:
            parts.append(f"| `{el}` | {kor or ''} | {ty or ''} | {req or ''} | {ln or ''} | {ds or ''} |")
        parts.append("")

    if api["request_example"]:
        parts.append("**Request Example:**")
        parts.append("```")
        parts.append(api["request_example"])
        parts.append("```")
        parts.append("")
    if api["response_example"]:
        parts.append("**Response Example:**")
        parts.append("```")
        parts.append(api["response_example"][:3000])  # 너무 긴 응답 예시는 잘라냄
        parts.append("```")
        parts.append("")

    parts.append("---")
    parts.append("")
    return "\n".join(parts)


def render_index(by_file: dict[str, list[dict]]) -> str:
    """전체 API 목록 + 검색 가능한 인덱스."""
    lines: list[str] = []
    lines.append("# 한국투자증권 OpenAPI — Index")
    lines.append("")
    lines.append("339개 API의 LLM 친화 변환본. 원본은 `한국투자증권_오픈API_전체문서_*.xlsx` (소스 보존).")
    lines.append("")
    lines.append("## 카테고리별 파일")
    lines.append("")
    lines.append("| 파일 | 카테고리 | API 수 |")
    lines.append("|---|---|---|")
    total = 0
    for fname, apis in sorted(by_file.items()):
        cat = apis[0].get("_category_label", "")
        lines.append(f"| [{fname}](./{fname}) | {cat} | {len(apis)} |")
        total += len(apis)
    lines.append(f"| **TOTAL** | | **{total}** |")
    lines.append("")

    lines.append("## 공통 인증 헤더 (모든 REST API 공통, 각 API 본문에서 생략됨)")
    lines.append("")
    lines.append("| Element | 의미 | 비고 |")
    lines.append("|---|---|---|")
    lines.append("| `content-type` | application/json; charset=utf-8 | 고정 |")
    lines.append("| `authorization` | `Bearer {access_token}` | OAuth 토큰. `KisClient`가 자동 부여 |")
    lines.append("| `appkey` / `appsecret` | 한투 발급 키 | Workers Secrets |")
    lines.append("| `tr_id` | 거래 ID | API별 상이 (각 섹션 메타 표 참조) |")
    lines.append("| `tr_cont` | 연속 조회 키 | 페이징 응답이 N건 cap일 때 |")
    lines.append("| `custtype` | `P`(개인) / `B`(법인) | 기본 P |")
    lines.append("")

    lines.append("## 전체 API 목록 (TR_ID 검색용)")
    lines.append("")
    lines.append("| TR_ID | API 이름 | 카테고리 | URL |")
    lines.append("|---|---|---|---|")
    all_apis: list[tuple[str, str, str, str, str]] = []
    for fname, apis in by_file.items():
        for api in apis:
            m = api["meta"]
            tr = m.get("tr_real", "").strip() or "—"
            name = m.get("name") or m.get("sheet", "")
            cat = api.get("_category_label", "")
            url = m.get("url", "").strip() or "—"
            all_apis.append((tr, name, cat, url, fname))
    # 정렬: 카테고리 → TR_ID
    all_apis.sort(key=lambda x: (x[2], x[0]))
    for tr, name, cat, url, fname in all_apis:
        link = f"[{name}](./{fname}#{slug(name)})"
        lines.append(f"| `{tr}` | {link} | {cat} | `{url}` |")
    lines.append("")

    lines.append("## 본 MCP에서 사용 중인 핵심 API")
    lines.append("")
    lines.append("| 도구 | TR_ID | 시트명 |")
    lines.append("|---|---|---|")
    lines.append("| `get_quote` (주식) | `FHKST01010100` | 주식현재가 시세 |")
    lines.append("| `get_quote` (ETF/ETN) | `FHPST02400000` | ETF/ETN 현재가 |")
    lines.append("| `get_chart` | `FHKST03010100` | 국내주식기간별시세(일/주/월/년) |")
    lines.append("| `get_etf_components` | `FHKST121600C0` | ETF 구성종목시세 |")
    lines.append("| `get_dividend` | `HHKDB669102C0` | 예탁원정보(배당일정) |")
    lines.append("| `get_credit_ratio` | `FHPST04760000` | 종목별 일별 신용잔고 |")
    lines.append("| `get_credit_ratio` (공매도) | `FHPST04830000` | 종목별 일별 공매도 |")
    lines.append("| `get_credit_ratio` (대차) | `HHPST074500C0` | 종목별 일별 대차거래추이 |")
    lines.append("| `advanced_search` (등락률) | `FHPST01700000` | 국내주식 등락률 순위 |")
    lines.append("| `advanced_search` (시총) | `FHPST01740000` | 국내주식 시가총액 순위 |")
    lines.append("| `advanced_search` (거래량) | `FHPST01710000` | 거래량 순위 |")
    lines.append("| `get_index` (국내) | `FHPUP02100000` | 국내업종 현재지수 |")
    lines.append("| `get_index_chart` (국내) | `FHPUP02120000` | 국내업종 일자별지수 |")
    lines.append("| `get_index`/`get_fx` (해외) | `FHKST03030100` | 해외주식 종목/지수/환율 기간별시세 |")
    lines.append("| `get_commodity` | `HHDFC55010000` | 해외선물종목현재가 |")
    lines.append("| `get_commodity_chart` | `HHDFC55020100` | 해외선물 체결추이(일간) |")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: source xlsx not found: {XLSX}", file=sys.stderr)
        return 1

    print(f"Loading {XLSX.name} …")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    print(f"  {len(wb.sheetnames)} sheets")

    by_file: dict[str, list[dict]] = {}
    for name in wb.sheetnames:
        sheet = wb[name]
        try:
            api = extract_sheet(name, sheet)
        except Exception as e:
            print(f"  ! {name}: {e}", file=sys.stderr)
            continue
        fname, label = categorize(api["meta"])
        api["_category_label"] = label
        by_file.setdefault(fname, []).append(api)

    OUT.mkdir(parents=True, exist_ok=True)

    # 카테고리별 파일
    for fname, apis in by_file.items():
        out = OUT / fname
        header = [
            f"# {apis[0]['_category_label']} API",
            "",
            f"한국투자증권 OpenAPI — `{apis[0]['_category_label']}` 카테고리 ({len(apis)}개).",
            "원본 시트는 cp949 엑셀이며 본 파일은 LLM 친화 변환본. 검색은 `INDEX.md` 권장.",
            "",
            "공통 OAuth 헤더(`authorization`, `appkey`, `appsecret`, `tr_id`, `custtype` 등)는 모든 API 동일하므로 본 문서에서 생략. `INDEX.md` 상단 참고.",
            "",
            "---",
            "",
        ]
        body = "".join(render_section(api, i + 1) for i, api in enumerate(apis))
        out.write_text("\n".join(header) + body, encoding="utf-8")
        print(f"  wrote {out.relative_to(ROOT)} ({len(apis)} APIs)")

    # INDEX
    index_path = OUT / "INDEX.md"
    index_path.write_text(render_index(by_file), encoding="utf-8")
    print(f"  wrote {index_path.relative_to(ROOT)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
